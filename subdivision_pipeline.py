import torch
import math
import numpy as np
import pyvista as pv
import matplotlib.pyplot as plt
import torch.nn.functional as F
import time
import csv
import os
from openpyxl import Workbook, load_workbook

from help_functions.make_conv_3d_torch import make_conv_3d_torch
from help_functions.morph_fill import morph_fill_fast
from help_functions.plot_subdivision_points import plot_subdivision_points
from help_functions.shadows_visual import shadows_visual, points_visual
from help_functions.fast_optimization import (Sub_A_fast, 
                                              solve_least_squares_subdivision_CG,
                                              build_kernel)
from help_functions.spherical_kernel_3d import spherical_kernel_3d
from help_functions.make_html_visual import make_html_visual
from help_functions.paddings import pad_to_multiple_centered

root_folder = 'subdivision_results'
excel_path = os.path.join(root_folder, "timings.xlsx")
timings = {}
mask = torch.tensor([1 / 8, 4 / 8, 6 / 8, 4 / 8, 1 / 8], dtype=torch.float32)
j = 2
gd_iter = 3

#HAND
subfolder = 'hand'
camera_position = (750, 400,350)
slice_coord = 190
input_koef = 100
model_path = './models/HandWithNormals.txt'
rotations = {(2,0): 1,
             }

# #BUDDA
# subfolder = 'budda'
# camera_position = (550,350,250)
# slice_coord = 400
# input_koef = 3500
# model_path ='./models/BuddaAll.txt'
# rotations = {(1,2): 1,
#              }

# #PLANK
# subfolder = 'plank'
# camera_position = (750,200,150)
# slice_coord = 180
# input_koef = 1
# model_path ='./models/MaxPlankWithNormals.xyz'
# rotations = {(1,2): 1,
#              (0,1): 3,
#              }

# #TURBINE
# subfolder = 'turbine'
# camera_position = (750,-400,350)
# slice_coord = 135
# input_koef = 5
# model_path = './models/Turbine.txt'
# rotations = {(1,2): 1,
#              }

#ИМПОРТИРУЕМ И ПРЕДОБРАБАТЫВАЕМ ДАННЫЕ
print("ИМПОРТИРУЕМ ДАТАСЕТ")
t0 = time.perf_counter()

raw_data = np.loadtxt(model_path)[:, :3].reshape(-1)
print("Количество точек во входном файле: ", len(raw_data))
int_cast_data = np.trunc(raw_data*input_koef).astype(int).reshape(-1, 3)
shifted = (int_cast_data - int_cast_data.min(axis=0) + np.array([10,10,10]))
sizes = (int_cast_data.max(axis=0) - int_cast_data.min(axis=0) + np.array([20,20, 20])).astype(int)

model_tensor_3D = torch.zeros(tuple(sizes), dtype=torch.float16)
model_tensor_3D[shifted[:, 0], shifted[:, 1], shifted[:, 2]] = 1

timings["Import data"] = round(time.perf_counter() - t0, 2)


#При необходимости поворачиваем
model_tensor_3D_rotated = model_tensor_3D
for dims, k in rotations.items():
    model_tensor_3D_rotated = torch.rot90(input = model_tensor_3D_rotated, k=k, dims=dims)

point_count = len(torch.argwhere(model_tensor_3D != 0))
print (f"Количество точек {point_count}")
model_tensor_3D = model_tensor_3D_rotated

# timings["Points count"] = point_count


#УТОЛЩАЕМ ГРАНИЦЫ
print("УТОЛЩАЕМ ГРАНИЦЫ")
t0 = time.perf_counter()

radius = 3
kernel_3D = spherical_kernel_3d(radius)
model_conv_result_3D = make_conv_3d_torch(source=model_tensor_3D, kernel=kernel_3D)

timings["Borders convolution"] = round(time.perf_counter() - t0, 2)



#ВЫПОЛНЯЕМ МОРФОЛОГИЧЕСКОЕ ЗАПОЛНЕНИЕ
print("ВЫПОЛНЯЕМ МОРФОЛОГИЧЕСКОЕ ЗАПОЛНЕНИЕ")
t0 = time.perf_counter()

model_conv_result_3D_filled = model_conv_result_3D.clone()
for layer in range (model_conv_result_3D.shape[2]):
    model_conv_result_3D_filled[:,:,layer] = morph_fill_fast(model_conv_result_3D_filled[:,:,layer])

timings["Morph fill"] = round(time.perf_counter() - t0, 2)



#НАХОДИМ НАЧАЛЬННУЮ ПОСЛЕДОВАТЕЛЬНОСТЬ
print("НАХОДИМ НАЧАЛЬННУЮ ПОСЛЕДОВАТЕЛЬНОСТЬ")
t0 = time.perf_counter()

Z = model_conv_result_3D_filled
Z_pad, pads = pad_to_multiple_centered(Z, j)
x0 = solve_least_squares_subdivision_CG(Z, mask, j = j, tol = 1e-1, max_iter=gd_iter)

timings["Finding x0"] = round(time.perf_counter() - t0, 2)


#SUBDIVISION
print("ВЫПОЛНЯЕМ ИТЕРАЦИИ ПОДРАЗДЕЛЕНИЙ")
t0 = time.perf_counter()

kernel = build_kernel(mask, dim = 3)
x0_subdivisioned = x0.clone()
for i in range(j):
    x0_subdivisioned = Sub_A_fast(x0_subdivisioned, kernel)
    
timings["Subdivision"] = round(time.perf_counter() - t0, 2)


time_keys = [
    "Import data",
    "Borders convolution",
    "Morph fill",
    "Finding x0",
    "Subdivision"
]

total_time = sum(timings[k] for k in time_keys if k in timings)
timings["Total time"] = round(total_time, 2)

experiment_name = f"{subfolder}_{point_count}_points"
#СОХРАНЕНИЕ ОТЧЕТА В ТАБЛИЦУ
if os.path.exists(excel_path):
    wb = load_workbook(excel_path)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="stage")

# ищем, есть ли уже такой subfolder (столбец)
col = None
for c in range(2, ws.max_column + 1):
    if ws.cell(row=1, column=c).value == experiment_name:
        col = c
        break

# если нет — добавляем новый столбец
if col is None:
    col = ws.max_column + 1
    ws.cell(row=1, column=col, value=experiment_name)

row_map = {}

for r in range(2, ws.max_row + 1):
    stage_name = ws.cell(row=r, column=1).value
    row_map[stage_name] = r

for stage, value in timings.items():
    if stage in row_map:
        r = row_map[stage]
    else:
        r = ws.max_row + 1
        ws.cell(row=r, column=1, value=stage)
        row_map[stage] = r

    ws.cell(row=r, column=col, value=value)

wb.save(excel_path)

print("Сохранено в", excel_path)